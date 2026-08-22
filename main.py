import re
import pandas as pd
from datetime import date, datetime
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

import processor
from config import RUTA_MAESTRO, RUTA_DISPENSACION, RUTA_REMISIONES, RUTA_STOCK_BODEGA, RUTA_STOCK_PUNTOS, RUTA_MOLECULAS
from processor import ProyeccionProcessor

_CHARS_ILEGALES = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F]')

# Fila de procedencia: de donde sale cada columna. Se escribe encima de los
# titulos para que cualquiera entienda el origen del dato sin preguntar.
DESCRIPCIONES_COLUMNAS = {
    # --- identificacion (archivo Maestro / BD) ---
    'Codigo': 'Maestro (BD) - codigo del articulo',
    'Nombre': 'Maestro (BD) - nombre del articulo',
    'Grupo': 'Maestro (BD) - laboratorio',
    'Proveedor': 'Maestro (BD) - ultimo proveedor de compra',
    'Ultimo Costo': 'Maestro (BD) - ultimo costo de compra',
    'Costo Promedio': 'Maestro (BD) - costo promedio',
    'Codigo_Molecula': 'Archivo Molecula_Compra - codigo de molecula',
    'Molecula': 'Archivo Molecula_Compra - nombre de molecula',
    'N_Productos': 'Calculado - numero de productos en la molecula',
    # --- consumo por canal / contrato ---
    'Consumo_NEPS_Capita': 'Dispensacion - cliente NUEVA EPS + servicio CAPITADO',
    'Consumo_NEPS_Evento': 'Dispensacion - cliente NUEVA EPS + servicio POS / MIPRES / TUTELA',
    'Consumo_FOMAG_Evento': 'Dispensacion - cliente Fiduprevisora / FIDEICOMISOS (FOMAG)',
    'Consumo_Sin_Clasificar': 'Dispensacion - filas que no coincidieron con ninguna regla (deberia ser 0)',
    'Consumo_Remisiones': 'Archivo Remisiones - consumo total del periodo',
    'Consumo_Dispensacion_Total': 'Calculado - suma de los canales de dispensacion (Capita + Evento + FOMAG)',
    'Consumo_Molecula': 'Calculado - suma del Consumo_Acum de todos los productos de la molecula',
    'Consumo_Acum': 'Calculado - consumo total = dispensacion + remisiones (meses alineados)',
    # --- rotacion ---
    'Rotacion': 'Calculado - Pareto ABC sobre Consumo_Acum (A=80%, M=95%, B=resto)',
    'Rotacion_Dispensacion': 'Calculado - Pareto ABC solo sobre el consumo de dispensacion',
    'Rotacion_Remisiones': 'Calculado - Pareto ABC solo sobre el consumo de remisiones',
    # --- stock ---
    'Stock_Bodega_Principal': 'Valorizado CEDI - unidades en bodega principal',
    'Stock_Puntos_Dispensacion': 'Valorizado Puntos - suma de unidades de los puntos de dispensacion',
    'Stock_Total': 'Calculado - Stock_Bodega_Principal + Stock_Puntos_Dispensacion',
    'Bodega_Disponible_Comercial': 'Calculado - bodega (CEDI) que le queda a comercial tras dispensacion. El pedido comercial se resta contra esto: Necesidad_Rem - esto = Pedir_Remisiones',
    # --- motor de pedido ---
    'Demanda_Mensual': 'Calculado - demanda mensual ponderada 70/30 (ult. 3 meses 70%, primeros 3 meses 30%)',
    'Demanda_Disp_Mensual': 'Calculado - demanda mensual 70/30 solo del canal dispensacion',
    'Demanda_Rem_Mensual': 'Calculado - demanda mensual 70/30 solo del canal remisiones',
    'Demanda_Diaria': 'Calculado - Demanda_Mensual / 30 (consumo promedio por dia)',
    'Cobertura_Dias': 'Parametro (config.py) - dias de cobertura segun la rotacion',
    'Stock_Seguridad': 'Calculado - demanda diaria x dias de seguridad',
    'Necesidad_Disp': 'Calculado - unidades objetivo del canal dispensacion (demanda_disp/30 x factor_disp, dias de config *_DISP)',
    'Necesidad_Rem': 'Calculado - unidades objetivo del canal remisiones (demanda_rem/30 x factor_remi, dias de config *_REMI)',
    'Necesidad_Mensual': 'Calculado - Necesidad_Disp + Necesidad_Rem (necesidad total del producto)',
    'Cantidad a Pedir': 'Calculado - pedido total (modelo vigente) = Pedir_Dispensacion_Total + Pedir_Remisiones. La bodega la agota Dispensacion y Comercial toma el sobrante (sin doble resta)',
    'Pedir_Dispensacion_Total': 'Calculado - pedido del canal dispensacion vs stock total (bodega + puntos)',
    'Pedir_NEPS_Capita': 'Calculado - parte del pedido de dispensacion segun peso historico de Capita',
    'Pedir_NEPS_Evento': 'Calculado - parte del pedido de dispensacion segun peso historico de Evento',
    'Pedir_FOMAG_Evento': 'Calculado - parte del pedido de dispensacion segun peso historico de FOMAG',
    'Pedir_Remisiones': 'Calculado - pedido del canal comercial/remisiones (modelo vigente): objetivo de remisiones menos la bodega que le queda tras dispensacion. Es la parte comercial de Cantidad a Pedir (sin doble resta)',
    'Valorizado': 'Calculado - cantidad a pedir de ESTA hoja * Ultimo Costo (si es 0, usa Costo Promedio)',
    'Estado': 'Calculado - COMPRAR si hay cantidad a pedir en ESTA hoja, si no NO COMPRAR',
    'Sobrantes Disp': 'Sobrante (Objetivo disp - Stock_Total) (positivo = falta, negativo = sobra)',
    'Sobrantes Remi': 'Sobrante (objetivo remi - Bodega Sobrante) (positivo = falta, negativo = sobra)',
    'Total Sobrantes': 'Unidades físicas que sobran (sin doble conteo)'
}


def _descripcion_columna(col):
    """Procedencia de una columna. Las columnas mensuales se resuelven por patron."""
    if col in DESCRIPCIONES_COLUMNAS:
        return DESCRIPCIONES_COLUMNAS[col]
    col = str(col)
    if re.match(r'Consumo_Disp_[A-Za-z]{3}_\d{4}$', col):
        return 'Dispensacion - consumo del mes'
    if re.match(r'Consumo_Rem_[A-Za-z]{3}_\d{4}$', col):
        return 'Remisiones - consumo del mes'
    if re.match(r'Consumo_[A-Za-z]{3}_\d{4}$', col):
        return 'Dispensacion + Remisiones - consumo del mes'
    return ''


def _limpiar_df(df):
    """Quita caracteres de control que openpyxl rechaza."""
    df = df.copy()
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]) or pd.api.types.is_datetime64_any_dtype(df[col]):
            continue
        df[col] = df[col].astype(str).str.replace(_CHARS_ILEGALES, '', regex=True)
    return df

def leer_con_encabezado(ruta, col_clave, max_filas=10, **kwargs):
    filas_encabezado = None
    preview = pd.read_excel(ruta, header=None, nrows=max_filas)
    for i in range(len(preview)):
        valores_fila = preview.iloc[i].astype(str).str.strip().values
        if col_clave in valores_fila:
            filas_encabezado = i
            break
    if filas_encabezado is None:
        raise ValueError(f"No encontre '{col_clave}' en las primeras {max_filas} filas de {ruta}.")
    df = pd.read_excel(ruta, skiprows=filas_encabezado, **kwargs)

    if str(df.columns[0]).startswith('Unnamed') and df.iloc[:, 0].isna().all():
        df = df.drop(columns=df.columns[0])
    return df

def cargar_datos(RUTA_MAESTRO, RUTA_DISPENSACION, RUTA_REMISIONES, RUTA_STOCK_BODEGA,
                 RUTA_STOCK_PUNTOS, RUTA_MOLECULAS):
    maestro = leer_con_encabezado(RUTA_MAESTRO, 'Codigo')
    molecula_compra = leer_con_encabezado(RUTA_MOLECULAS, 'Codigo Articulo')
    consumo_dispensacion = pd.read_excel(RUTA_DISPENSACION)
    consumo_remisiones = pd.read_excel(RUTA_REMISIONES)
    stock_bodega = leer_con_encabezado(RUTA_STOCK_BODEGA, 'Codigo Articulo')
    stock_puntos = leer_con_encabezado(RUTA_STOCK_PUNTOS, 'Codigo Articulo')
    return maestro, consumo_dispensacion, consumo_remisiones, stock_bodega, stock_puntos, molecula_compra


def _valorizar_estado(sub, col_cantidad):

    sub['Valorizado'] = ProyeccionProcessor._valorizar(
        sub[col_cantidad], sub.get('Ultimo Costo'), sub.get('Costo Promedio'))
    sub['Estado'] = sub[col_cantidad].gt(0).map({True: 'COMPRAR', False: 'NO COMPRAR'})
    return sub


def construir_hojas(processor):

    df = processor.maestro_consumo
    mensuales = list(getattr(processor, 'cols_consumo_mensual', []))
    mensuales_disp = list(getattr(processor, 'cols_consumo_mensual_disp', []))
    mensuales_rem = list(getattr(processor, 'cols_consumo_mensual_rem', []))
    pedir_contratos = list(processor.SEGMENTOS_CONTRATO.values())



    hojas = {}

    # ---------------------------------------------------------- DISPENSACION

    colsDisp =['Codigo', 'Nombre', 'Codigo_Molecula', 'Molecula','Nombre Comercial', 'Grupo', 'Proveedor',
            'Costo Promedio', 'Ultimo Costo', 'Consumo_NEPS_Capita','Consumo_NEPS_Evento',
            'Consumo_FOMAG_Evento', 'Consumo_Sin_Clasificar', 'Consumo_Dispensacion_Total',
            *mensuales_disp, 'Demanda_Disp_Mensual', 'Rotacion_Dispensacion', 'Stock_Bodega_Principal',
            'Stock_Puntos_Dispensacion', 'Stock_Total', 'Necesidad_Disp', 'Pedir_NEPS_Capita', 'Pedir_NEPS_Evento',
            'Pedir_FOMAG_Evento', 'Pedir_Dispensacion_Total', 'Estado', 'Valorizado Ult Costo', 'Valorizado Promedio']
    mask = (df.get('Consumo_Dispensacion_Total', 0) > 0) | (df.get('Pedir_Dispensacion_Total', 0) > 0)
    disp = (df.loc[mask, [c for c in colsDisp if c in df.columns]]
            .sort_values('Pedir_Dispensacion_Total', ascending=False).copy())
    _valorizar_estado(disp, 'Pedir_Dispensacion_Total')   # valorizado del canal dispensacion
    hojas['Dispensacion'] = disp

    # ------------------------------------------------------------ REMISIONES
    # Remisiones son clientes distintos: el stock de puntos NO les aplica (ni se
    # muestra ni interviene en el pedido). Solo bodega principal.
    colsRem = ['Codigo', 'Nombre', 'Codigo_Molecula', 'Molecula','Nombre Comercial', 'Grupo', 'Proveedor',
               'Costo Promedio', 'Ultimo Costo', *mensuales_rem, 'Consumo_Remisiones','Demanda_Rem_Mensual',
               'Rotacion_Remisiones','Stock_Bodega_Principal', 'Bodega_Disponible_Comercial', 'Necesidad_Rem','Pedir_Remisiones',
               'Estado', 'Valorizado Ult Costo', 'Valorizado Promedio']
    mask = (df.get('Consumo_Remisiones', 0) > 0) | (df.get('Pedir_Remisiones', 0) > 0)
    rem = (df.loc[mask, [c for c in colsRem if c in df.columns]]
           .sort_values('Pedir_Remisiones', ascending=False).copy())
    _valorizar_estado(rem, 'Pedir_Remisiones')            # valorizado del canal remisiones
    hojas['Remisiones'] = rem

    # ------------------------------------------------------------------ TODO
    colsTodo = ['Codigo', 'Nombre','Presentación', 'Codigo_Molecula', 'Molecula','Nombre Comercial', 'Grupo', 'Proveedor',
                'Costo Promedio', 'Ultimo Costo', 'Consumo_Molecula', 'Consumo_NEPS_Capita',
                'Consumo_NEPS_Evento', 'Consumo_FOMAG_Evento', 'Consumo_Dispensacion_Total',
                'Consumo_Remisiones', *mensuales, 'Consumo_Acum',
                'Rotacion', 'Rotacion_Dispensacion', 'Rotacion_Remisiones','Demanda_Disp_Mensual',
                'Demanda_Rem_Mensual','Demanda_Mensual','Demanda_Diaria',
                'Necesidad_Disp','Necesidad_Rem','Necesidad_Mensual','Valorizado Promedio Sin Rest Inv','Valorizado Ult Costo Sin Rest Inv', 'Stock_Bodega_Principal',
                'Stock_Puntos_Dispensacion','Stock_Total','Sobrantes Disp','Sobrantes Remi','Total Sobrantes','Cantidad a Pedir','Estado',
                'Valorizado Promedio','Valorizado Ult Costo','Pedir_NEPS_Capita', 'Pedir_NEPS_Evento','Pedir_FOMAG_Evento',
                'Pedir_Dispensacion_Total','Pedir_Remisiones']
    hojas['Todo'] = (df[[c for c in colsTodo if c in df.columns]]
                     .sort_values('Cantidad a Pedir', ascending=False))

    # -------------------------------------------------------------- TODO SIN CEDI
    colsTodoSinCedi = ['Codigo', 'Nombre','Presentación', 'Codigo_Molecula', 'Molecula','Nombre Comercial', 'Grupo', 'Proveedor',
                'Costo Promedio', 'Ultimo Costo', 'Consumo_Molecula', 'Consumo_NEPS_Capita',
                'Consumo_NEPS_Evento', 'Consumo_FOMAG_Evento', 'Consumo_Dispensacion_Total',
                'Consumo_Remisiones', *mensuales, 'Consumo_Acum',
                'Rotacion', 'Rotacion_Dispensacion', 'Rotacion_Remisiones','Demanda_Disp_Mensual',
                'Demanda_Rem_Mensual','Demanda_Mensual','Demanda_Diaria',
                'Necesidad_Disp','Necesidad_Rem','Necesidad_Mensual','Valorizado Promedio Sin Rest Inv','Valorizado Ult Costo Sin Rest Inv', 'Stock_Bodega_Principal',
                'Stock_Puntos_Dispensacion','Stock_Total','Sobrantes Disp sin CEDI','Sobrantes Remi sin CEDI','Total Sobrantes sin CEDI','Cantidad a Pedir sin CEDI','Estado',
                'Valorizado Promedio sin CEDI','Valorizado Ult Costo sin CEDI','Pedir_NEPS_Capita sin CEDI', 'Pedir_NEPS_Evento sin CEDI','Pedir_FOMAG_Evento sin CEDI',
                'Pedir Dispensacion Total sin CEDI','Pedir Remisiones sin CEDI']
    hojas['Todo REST CEDI'] = (df[[c for c in colsTodoSinCedi if c in df.columns]]
                            .sort_values('Cantidad a Pedir sin CEDI', ascending=False)).copy()

    #--------------------------------------------------------------- TODO SIN PUNTOS
    colsTodoSinPuntos = ['Codigo', 'Nombre','Presentación', 'Codigo_Molecula', 'Molecula','Nombre Comercial', 'Grupo', 'Proveedor',
                'Costo Promedio', 'Ultimo Costo', 'Consumo_Molecula', 'Consumo_NEPS_Capita',
                'Consumo_NEPS_Evento', 'Consumo_FOMAG_Evento', 'Consumo_Dispensacion_Total',
                'Consumo_Remisiones', *mensuales, 'Consumo_Acum',
                'Rotacion', 'Rotacion_Dispensacion', 'Rotacion_Remisiones','Demanda_Disp_Mensual',
                'Demanda_Rem_Mensual','Demanda_Mensual','Demanda_Diaria',
                'Necesidad_Disp','Necesidad_Rem','Necesidad_Mensual','Valorizado Promedio Sin Rest Inv','Valorizado Ult Costo Sin Rest Inv', 'Stock_Bodega_Principal',
                'Stock_Puntos_Dispensacion','Stock_Total','Sobrantes Disp sin PUNTOS','Sobrantes Remi sin PUNTOS','Total Sobrantes sin PUNTOS','Cantidad a Pedir sin PUNTOS','Estado',
                'Valorizado Promedio sin PUNTOS','Valorizado Ult Costo sin PUNTOS','Pedir_NEPS_Capita sin PUNTOS', 'Pedir_NEPS_Evento sin PUNTOS','Pedir_FOMAG_Evento sin PUNTOS',
                'Pedir Dispensacion Total sin PUNTOS','Pedir Remisiones sin PUNTOS']
    hojas['Todo REST PUNTOS'] = (df[[c for c in colsTodoSinPuntos if c in df.columns]]
    )                           .sort_values('Cantidad a Pedir sin PUNTOS', ascending=False).copy()

    # -------------------------------------------------------------- MOLECULA
    cols = ['Codigo_Molecula', 'Molecula','Nombre Comercial', 'N_Productos', 'Rotacion',
            'Consumo_Molecula', 'Stock_Total', 'Demanda_Mensual', 'Demanda_Diaria', 'Cobertura_Dias',
            'Stock_Seguridad', 'Necesidad_Mensual',
            'Cantidad a Pedir', 'Pedir_Dispensacion_Total', *pedir_contratos]
    mol = processor.pedido_molecula
    hojas['Pedido_Molecula'] = (mol[[c for c in cols if c in mol.columns]]
                                .sort_values('Cantidad a Pedir', ascending=False))

    for sub in hojas.values():
        sub['Usuario ODC'] = ''

    return hojas


def exportar_excel(processor, ruta=None, buffer=None):
    if ruta is None:
        ruta = f'Excel Proyeccion/Resultados_Proyeccion_{date.today():%Y-%m-%d}.xlsx'

    hojas = construir_hojas(processor)

    fuente_desc = Font(italic=True, size=8, color='555555')
    fuente_titulo = Font(bold=True)
    relleno_desc = PatternFill('solid', fgColor='FFF3E0')
    alineacion_desc = Alignment(wrap_text=True, vertical='top')
    grupo_1_color = PatternFill('solid', fgColor='D9EAD3')
    grupo_2_color = PatternFill('solid', fgColor='CCF2FF')
    color_fila_estado_rojo = PatternFill('solid', fgColor='FF9B9B')
    color_fila_estado_verde = PatternFill('solid', fgColor='D9EAD3')
    color_celda_estado = PatternFill('solid', fgColor='FFFF09')
    color_celdas_sobrantes = PatternFill('solid', fgColor='9EB8CA')


    def colorear_encabezados(ws, sub, nombre_col, fill):
        if nombre_col in sub.columns:

            idx = list(sub.columns).index(nombre_col) + 1
            ws.cell(row=2, column=idx).fill = fill

    def _escribir(destino):
        with pd.ExcelWriter(destino, engine='openpyxl') as writer:
            for nombre, sub in hojas.items():
                sub = _limpiar_df(sub)
                # startrow=1 deja libre la fila 1 para la procedencia;
                # los titulos quedan en la fila 2 y los datos desde la 3.
                sub.to_excel(writer, sheet_name=nombre[:31], index=False, startrow=1)
                ws = writer.sheets[nombre[:31]]

                if 'Usuario ODC' in sub.columns:
                    dv = DataValidation(
                        type="list",
                        formula1='"NMEDINA,DGUACAS,YPEREZ,DMENESES,DHURTADO,JMONCAYO,DISPONIBLE OTRO CODIGO,EN BODEGA,'
                                 'DESCONTINUADO,AGOTADO,BAJO PEDIDO,BAJA ROTACION,SOLICITAR OTRO PROVEEDOR"',
                        allow_blank=True
                    )
                    ws.add_data_validation(dv)

                    idx = list(sub.columns).index('Usuario ODC') + 1
                    letra = ws.cell(row=2, column=idx).column_letter
                    dv.add(f'{letra}3:{letra}{ws.max_row}')

                for j, col in enumerate(sub.columns, start=1):
                    celda = ws.cell(row=1, column=j, value=_descripcion_columna(col))
                    celda.font = fuente_desc
                    celda.fill = relleno_desc
                    celda.alignment = alineacion_desc
                    ws.cell(row=2, column=j).font = fuente_titulo

                ws.freeze_panes = 'A3'                       # fija procedencia + titulos
                ws.row_dimensions[1].height = 60
                if ws.max_row > 2:
                    ws.auto_filter.ref = f'A2:{ws.cell(2, ws.max_column).coordinate}'

                if nombre in ('Todo', 'Todo REST CEDI', 'Todo REST PUNTOS') and 'Estado' in sub.columns:
                    col_estado = list(sub.columns).index('Estado') + 1
                    for r in range(3, ws.max_row + 1):
                        celda_estado = ws.cell(row=r, column=col_estado)
                        valor = str(celda_estado.value).strip().upper() if celda_estado.value is not None else ""
                        if valor == 'COMPRAR':
                            celda_estado.fill = color_fila_estado_verde
                        if valor == 'NO COMPRAR':
                            celda_estado.fill = color_fila_estado_rojo
                    for c in ['Sobrantes Disp', 'Sobrantes Remi', 'Total Sobrantes']:
                        colorear_encabezados(ws, sub, c, color_celdas_sobrantes)
                    for c in ['Sobrantes Disp sin CEDI', 'Sobrantes Remi sin CEDI', 'Total Sobrantes sin CEDI']:
                        colorear_encabezados(ws, sub, c, color_celdas_sobrantes)
                    for c in ['Sobrantes Disp sin PUNTOS', 'Sobrantes Remi sin PUNTOS', 'Total Sobrantes sin PUNTOS']:
                        colorear_encabezados(ws, sub, c, color_celdas_sobrantes)
                    for e in ('Estado','Usuario ODC'):
                        colorear_encabezados(ws, sub, e, color_celda_estado)
                    for j in ('Cantidad a Pedir', 'Valorizado Promedio', 'Valorizado Ult Costo'):
                        colorear_encabezados(ws, sub, j, grupo_1_color)
                    for i in ('Cantidad a Pedir sin CEDI', 'Valorizado Promedio sin CEDI', 'Valorizado Ult Costo sin CEDI'):
                        colorear_encabezados(ws, sub, i, grupo_1_color)
                    for h in ('Cantidad a Pedir sin PUNTOS', 'Valorizado Promedio sin PUNTOS', 'Valorizado Ult Costo sin PUNTOS'):
                        colorear_encabezados(ws, sub, h, grupo_1_color)
                    for c in ('Necesidad_Mensual', 'Valorizado Promedio Sin Rest Inv', 'Valorizado Ult Costo Sin Rest Inv'):
                        colorear_encabezados(ws, sub, c, grupo_2_color)

                # Ancho por columna, ignorando la fila de procedencia (es larga y
                # va con ajuste de texto); se mide sobre titulo + una muestra de datos.
                for j in range(1, ws.max_column + 1):
                    letra = ws.cell(row=2, column=j).column_letter
                    muestra = [ws.cell(row=r, column=j).value
                               for r in range(2, min(ws.max_row, 300) + 1)]
                    ancho = max((len(str(v)) for v in muestra if v is not None), default=8)
                    ws.column_dimensions[letra].width = min(max(ancho + 2, 12), 40)

    if buffer is not None:
        _escribir(buffer)
        buffer.seek(0)
        return buffer

    if ruta is None:
        ruta = f'Excel Proyeccion/Resultados_Proyeccion_{date.today():%Y-%m-%d}.xlsx'
    try:
        _escribir(ruta)
    except PermissionError:
        # El archivo queda bloqueado si esta abierto en Excel
        alterna = f'Resultados_Proyeccion_{date.today():%Y-%m-%d}_{datetime.now():%H%M%S}.xlsx'
        print(f"\n[!] '{ruta}' esta abierto o bloqueado. Escribiendo en '{alterna}'.")
        _escribir(alterna)
        ruta = alterna

    print(f'\n[OK] Exportado: {ruta}')
    for nombre, sub in hojas.items():
        print(f'       {nombre:<18} {len(sub):>6,} filas x {len(sub.columns):>2} columnas')
    return ruta


def main():
    maestro, consumo_dispensacion, consumo_remisiones, stock_bodega, stock_puntos, molecula_compra = cargar_datos(RUTA_MAESTRO, RUTA_DISPENSACION, RUTA_REMISIONES,
                                                                                                                  RUTA_STOCK_BODEGA, RUTA_STOCK_PUNTOS, RUTA_MOLECULAS)

    processor = ProyeccionProcessor(maestro, consumo_dispensacion, consumo_remisiones, stock_bodega, stock_puntos, molecula_compra)
    processor.procesar()

    processor.auditoria_integridad()
    processor.imprimir_resumen_contratos()
    exportar_excel(processor)

    df = processor.maestro_consumo
    cantPas = df['Cant. Pastillas']

    a_m_disp = df['Rotacion_Dispensacion'].isin(['A', 'M']) & (cantPas >= 2)
    malos_disp = df[a_m_disp & (df['Pedir_Dispensacion_Total'] % cantPas != 0)]
    print('A/M disp fuera del multiplo: ', len(malos_disp))

    a_m_rem = df['Rotacion_Remisiones'].isin(['A', 'M']) & (cantPas >= 2)
    malos_rem = df[a_m_rem & (df['Pedir_Remisiones'] % cantPas != 0)]
    print('A/M rem fuera del multiplo: ', len(malos_rem))

    for col, rot in [('Pedir_Dispensacion_Total', 'Rotacion_Dispensacion'),
                     ('Pedir_Remisiones', 'Rotacion_Remisiones')]:
        rot_B = (df[rot] == 'B') & (cantPas >= 2)

        fuera = df[rot_B & (df[col] % cantPas != 0)]
        print(f'B {col} fuera del multiplo: ', len(fuera))


if __name__ == '__main__':
    main()
